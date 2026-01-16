
import openpyxl
import sys
import os

file_path = r"c:\Users\gonza\Desktop\PricingEngine\public\Dashboard_Report_2026-01-16 (15).xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"Successfully loaded {file_path}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")
        
        # Print first few rows found
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        for idx, row in enumerate(rows):
            print(f"Row {idx+1}: {row}")

        # Check for charts
        if hasattr(ws, '_charts') and ws._charts:
             print(f"Charts found: {len(ws._charts)}")
        else:
             print("No charts found (or cannot access via private attribute)")

except Exception as e:
    print(f"Error reading excel: {e}")
