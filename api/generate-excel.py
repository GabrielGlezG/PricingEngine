# Vercel Python Serverless Function for Excel Chart Generation
# Path: api/generate-excel.py

from http.server import BaseHTTPRequestHandler
import json
import io
from datetime import datetime, timedelta

# openpyxl imports
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.title import Title
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont, RegularTextRun, LineBreak
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64


# Styling constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Avenir Medium", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="Avenir Medium", size=10)
ALT_ROW_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL
            cell.border = THIN_BORDER


def apply_chart_styling(chart):
    """
    Attempts to enforce Avenir Font on Chart Elements using RichText.
    This works by replacing the standard Text object with a RichText object defined with specific font properties.
    """
    try:
        # Define Avenir Font Properties
        # sz is 100ths of point, so 1100 = 11pt
        font_title = DrawingFont(typeface='Avenir Black')
        cp_title = CharacterProperties(latin=font_title, sz=1400, b=True) 
        pp_title = ParagraphProperties(defRPr=cp_title)
        
        font_axis = DrawingFont(typeface='Avenir Medium')
        cp_axis = CharacterProperties(latin=font_axis, sz=900)
        pp_axis = ParagraphProperties(defRPr=cp_axis)

        # 1. Chart Title
        if chart.title and isinstance(chart.title, str):
            run = RegularTextRun(t=chart.title, rPr=cp_title)
            rt_title = RichText(p=[Paragraph(pPr=pp_title, endParaRPr=cp_title, r=[run])])
            chart.title = Title(tx=rt_title)
            
        # 2. X-Axis Title & Ticks
        if chart.x_axis:
            # Title
            if chart.x_axis.title and isinstance(chart.x_axis.title, str):
                run_x = RegularTextRun(t=chart.x_axis.title, rPr=cp_axis)
                rt_x = RichText(p=[Paragraph(pPr=pp_axis, endParaRPr=cp_axis, r=[run_x])])
                chart.x_axis.title = Title(tx=rt_x)
            # Ticks (Numbers/Categories)
            # Ticks don't take text run (values are dynamic), but we set the Paragraph Props default
            chart.x_axis.textProperties = RichText(p=[Paragraph(pPr=pp_axis, endParaRPr=cp_axis)])

        # 3. Y-Axis Title & Ticks
        if chart.y_axis:
            # Title
            if chart.y_axis.title and isinstance(chart.y_axis.title, str):
                run_y = RegularTextRun(t=chart.y_axis.title, rPr=cp_axis)
                rt_y = RichText(p=[Paragraph(pPr=pp_axis, endParaRPr=cp_axis, r=[run_y])])
                chart.y_axis.title = Title(tx=rt_y)
            # Ticks (Numbers)
            chart.y_axis.textProperties = RichText(p=[Paragraph(pPr=pp_axis, endParaRPr=cp_axis)])
            
        # 4. Legend Font
        if chart.legend:
             chart.legend.textProperties = RichText(p=[Paragraph(pPr=pp_axis, endParaRPr=cp_axis)])

    except Exception as e:
        # Fail silently - do not crash generation just for font
        print(f"Font styling warning: {str(e)}")


def create_bar_chart(ws, title, data_range, start_row, num_series):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Valor"
    
    data = Reference(ws, min_col=2, min_row=start_row, max_col=1 + num_series, max_row=data_range)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_range)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 28
    chart.height = 17
    
    apply_chart_styling(chart)
    
    return chart


def create_stacked_chart(ws, title, data_range, start_row, num_series):
    """Stacked bar chart for composition/breakdown data"""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"  # STACKED instead of clustered
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Total"
    
    data = Reference(ws, min_col=2, min_row=start_row, max_col=1 + num_series, max_row=data_range)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_range)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 28
    chart.height = 17
    
    apply_chart_styling(chart)
    
    return chart


def create_line_chart(ws, title, data_range, start_row, num_series):
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Valor"
    # Ensure dates are at the very bottom, not crossing the negative values
    chart.x_axis.tickLblPos = "low"
    
    data = Reference(ws, min_col=2, min_row=start_row, max_col=1 + num_series, max_row=data_range)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_range)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 28
    chart.height = 17
    
    apply_chart_styling(chart)
    
    return chart


def create_scatter_chart(ws, title, data_range, start_row, num_series):
    """Create a bubble chart for Matriz Posicionamiento where size = volume"""
    from openpyxl.chart import BubbleChart, Series, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    
    chart = BubbleChart()
    chart.title = title
    chart.style = 10
    chart.x_axis.title = "Volumen"
    chart.y_axis.title = "Precio"
    chart.bubbleScale = 30 # Back to standard size
    
    # Disable Data Labels on the chart itself (per user request)
    chart.dataLabels = None
    
    # Enable Legend at the Bottom ("abajo del eje x con el color")
    chart.legend.position = 'b'

    # Iterate through each row of data to create a distinct Series
    # This allows us to label each bubble with its specific "Brand - Model" name
    # structure: Col 1=Name, Col 2=Volumen(X), Col 3=Precio(Y)
    for i in range(start_row + 1, data_range + 1):
        # Series Title from Column 1 (Brand - Model)
        title_val = ws.cell(row=i, column=1).value
        title_str = str(title_val) if title_val else "Series"
        
        # Values
        x_val = Reference(ws, min_col=2, min_row=i) # Volumen
        y_val = Reference(ws, min_col=3, min_row=i) # Precio
        z_val = Reference(ws, min_col=2, min_row=i) # Size = Volumen
        
        series = Series(values=y_val, xvalues=x_val, zvalues=z_val)
        
        # Fix: XYSeries.title requires SeriesLabel object, not string
        series.title = SeriesLabel(v=title_str)
        
        chart.series.append(series)
    
    chart.width = 28
    chart.height = 17 # Taller to accommodate legend at bottom
    
    apply_chart_styling(chart)
    
    return chart


def generate_error_excel(error_message):
    """Creates a simple Excel file with the error message"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ERROR REPORT"
    ws['A1'] = "Critical Error generating Excel"
    ws['A1'].font = Font(color="FF0000", bold=True, size=14)
    ws['A3'] = str(error_message)
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 80
    
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook.getvalue()

def set_global_styles(wb):
    """
    Modifies the default styles to ensure Avenir Medium is the baseline font.
    This helps chart elements that inherit 'Document Default' to use Avenir.
    """
    try:
        # Access the 'Normal' named style
        if 'Normal' in wb.named_styles:
            normal = wb.named_styles['Normal']
            normal.font = Font(name='Avenir Medium', size=11)
        
        # Also try to set it on the default style of the workbook logic
        # (OpenPyXL internal default)
        
    except Exception as e:
        print(f"Global style warning: {e}")

def enforce_global_font(wb):
    """
    Final pass: Iterates through EVERY cell in ALL sheets to enforce Avenir Medium.
    This ensures no cell is left behind with default settings.
    """
    try:
        avenir = Font(name="Avenir Medium", size=10) # Base font
        
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                 for cell in row:
                     if cell.font and cell.font.name != "Avenir Medium":
                         # Create new font preserving basics
                         current = cell.font
                         new_f = Font(
                             name="Avenir Medium",
                             size=current.size if current.size else 10,
                             bold=current.bold,
                             italic=current.italic,
                             color=current.color
                         )
                         cell.font = new_f
                     elif not cell.font:
                         cell.font = avenir
    except Exception as e:
        print(f"Enforce font error: {e}")

def generate_excel(data):
    # Initialize Debug Log
    debug_log = []
    
    try:
        wb = Workbook()
        
        # Apply Global Font Settings immediately
        set_global_styles(wb)

        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)
        
        # Debug Sheet (Hidden)
        debug_ws = wb.create_sheet("DEBUG LOG")
        debug_ws.sheet_state = 'hidden'
        
        sheets = data.get('sheets', [])
        summary = data.get('summary', None)
        models = data.get('models', None)
        currency_symbol = data.get('currencySymbol', '$')
        timezone_offset = data.get('timezoneOffset', -3)
        title = data.get('title', 'REPORTE DE DASHBOARD')
        filters_data = data.get('filters', {})
        
        local_time = datetime.utcnow() + timedelta(hours=timezone_offset)
        
        wb = Workbook()
        default_sheet = wb.active # Will be removed later
        
        # Summary Sheet (Dashboard style)
        if summary:
            ws = wb.create_sheet("Resumen Ejecutivo", 0)
            ws.sheet_view.showGridLines = False # White background
            ws['A1'] = title
            ws['A1'].font = Font(name="Avenir Black", bold=True, size=16)
            ws['A2'] = f"Generado: {local_time.strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name="Avenir Medium", italic=True, color="666666")
            
            ws['A4'] = "Métrica"
            ws['B4'] = "Valor"
            style_header_row(ws, 4, 2)
            
            metrics = [
                ("Total Modelos", summary.get('total_models', 0)),
                ("Total Marcas", summary.get('total_brands', 0)),
                ("Precio Promedio", summary.get('avg_price', 0)),
                ("Precio Mediano", summary.get('median_price', 0)),
                ("Precio Mínimo", summary.get('min_price', 0)),
                ("Precio Máximo", summary.get('max_price', 0)),
                ("Desviación Estándar", summary.get('price_std_dev', 0)),
                ("Coef. Variación", summary.get('variation_coefficient', 0)),
                ("Descuento Promedio", summary.get('avg_discount_pct', 0)), 
            ]
            
            for i, (label, value) in enumerate(metrics):
                row = 5 + i
                # Apply Avenir Font
                c1 = ws.cell(row=row, column=1, value=label)
                c1.font = Font(name="Avenir Medium", size=10)
                c2 = ws.cell(row=row, column=2, value=value)
                c2.font = Font(name="Avenir Medium", size=10)
                
                if 2 <= i <= 6:
                    c2.number_format = f'"{currency_symbol}" #,##0'
                elif i >= 7: 
                    c2.number_format = '0.00%'
            
            style_data_rows(ws, 5, 13, 2)
            
            ws['A14'] = "Filtros Aplicados"
            ws['A14'].font = Font(name="Avenir Black", bold=True)
            
            filters = summary.get('filters', {})
            ws['A15'] = "Segmento:"
            ws['B15'] = ', '.join(filters.get('tipoVehiculo', [])) or "Todos"
            ws['A16'] = "Marca:"
            ws['B16'] = ', '.join(filters.get('brand', [])) or "Todas"
            ws['A17'] = "Modelo:"
            ws['B17'] = ', '.join(filters.get('model', [])) or "Todos"
            
            # Apply font to filters
            for r in range(15, 18):
                ws[f'A{r}'].font = Font(name="Avenir Medium", bold=True)
                ws[f'B{r}'].font = Font(name="Avenir Medium")

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 30
        
        # Info sheet for generic exports
        elif title and not summary:
            ws = wb.create_sheet("Información", 0)
            ws.sheet_view.showGridLines = False # White background
            ws['A1'] = title
            ws['A1'].font = Font(name="Avenir Black", bold=True, size=16)
            ws['A2'] = f"Generado: {local_time.strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name="Avenir Medium", italic=True, color="666666")
            
            if filters_data:
                current_row = 4
                ws.cell(row=current_row, column=1, value="Filtros Aplicados").font = Font(name="Avenir Black", bold=True)
                current_row += 1
                
                for filter_name, filter_values in filters_data.items():
                    if isinstance(filter_values, list) and len(filter_values) > 0:
                        ws.cell(row=current_row, column=1, value=f"{filter_name}:").font = Font(name="Avenir Medium", bold=True)
                        ws.cell(row=current_row, column=2, value=', '.join(filter_values)).font = Font(name="Avenir Medium")
                        current_row += 1
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
        
        # Chart sheets created below...

        # Models sheet
        if models and len(models) > 0:
            ws = wb.create_sheet("Modelos")
            ws.sheet_view.showGridLines = False # White background
            headers = ["Marca", "Modelo", "Versión", "Estado", "Tipo Vehículo", 
                    "Precio c/Bono", "Precio Lista", "Bono", "% Descuento"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(name="Avenir Medium", bold=True, color="FFFFFF")
            
            style_header_row(ws, 1, len(headers))
            
            for row_idx, model in enumerate(models, 2):
                ws.cell(row=row_idx, column=1, value=model.get('brand', ''))
                ws.cell(row=row_idx, column=2, value=model.get('model', ''))
                ws.cell(row=row_idx, column=3, value=model.get('submodel', '-'))
                ws.cell(row=row_idx, column=4, value=model.get('estado', 'N/A'))
                ws.cell(row=row_idx, column=5, value=model.get('tipo_vehiculo', 'N/A'))
                
                precio_bono = model.get('precio_con_bono', 0)
                precio_lista = model.get('precio_lista', 0)
                bono = model.get('bono', 0)
                
                # Calculate Discount % (Bono / Lista)
                if precio_lista and precio_lista > 0:
                    diff = (bono / precio_lista)
                else:
                    diff = 0
                
                c6 = ws.cell(row=row_idx, column=6, value=precio_bono)
                c6.number_format = f'"{currency_symbol}" #,##0'
                c7 = ws.cell(row=row_idx, column=7, value=precio_lista)
                c7.number_format = f'"{currency_symbol}" #,##0'
                c8 = ws.cell(row=row_idx, column=8, value=bono)
                c8.number_format = f'"{currency_symbol}" #,##0'
                c9 = ws.cell(row=row_idx, column=9, value=diff)
                c9.number_format = '0.0%'
                
                # Apply data font
                for c in [1,2,3,4,5,6,7,8,9]:
                    ws.cell(row=row_idx, column=c).font = Font(name="Avenir Medium", size=10)
            
            end_row = len(models) + 1
            style_data_rows(ws, 2, end_row, len(headers))
            
            widths = [15, 20, 25, 12, 15, 18, 18, 15, 15]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
        
        # Chart sheets
        for sheet_data in sheets:
            try:
                sheet_name = sheet_data.get('name', 'Sheet')[:31]
                chart_type = sheet_data.get('chart_type', 'bar')
                chart_title = sheet_data.get('chart_title', sheet_name)
                rows = sheet_data.get('data', [])
                
                if not rows:
                    debug_log.append(f"Skipping {sheet_name}: No data rows")
                    continue
                
                # 1. Create Data Sheet
                ws = wb.create_sheet(sheet_name)
                ws.sheet_view.showGridLines = False # White background
                
                headers = list(rows[0].keys())
                num_cols = len(headers)
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(name="Avenir Medium", bold=True, color="FFFFFF") # Avenir Header
                
                style_header_row(ws, 1, num_cols)
                
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Apply Avenir Font to Data
                        cell.font = Font(name="Avenir Medium", size=10)
                        
                        if col_idx > 1 and isinstance(value, (int, float)):
                            header_lower = header.lower()
                            sheet_name_lower = sheet_name.lower()
                            
                            # 1. Percentage Rules
                            if (
                                'variacion' in header_lower or 
                                '%' in header_lower or 
                                'volatilidad' in sheet_name_lower or
                                'tendencia' in sheet_name_lower
                            ):
                                cell.number_format = '0.00%'
                            
                            # 2. Integer/Count Rules
                            elif (
                                'cantidad' in header_lower or 
                                'volumen' in header_lower or 
                                'versiones' in header_lower or
                                'count' in header_lower or
                                'numero' in header_lower or
                                'composición' in sheet_name_lower or
                                'composicion' in sheet_name_lower
                            ):
                                cell.number_format = '#,##0'
                            
                            # 3. Currency Rules
                            else:
                                cell.number_format = f'"{currency_symbol}" #,##0'
                
                end_row = len(rows) + 1
                style_data_rows(ws, 2, end_row, num_cols)
                
                for col in range(1, num_cols + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 18
                
                # 2. Create Chart (on Separate Sheet)
                # Name: "Gráfico {Name}" (Truncated to 31 chars)
                chart_sheet_name = f"Gráfico {sheet_name}"[:31]
                
                # Ensure unique name
                counter = 1
                while chart_sheet_name in wb.sheetnames:
                    chart_sheet_name = f"Gráfico {sheet_name[:20]} {counter}"
                    counter += 1

                ws_chart = wb.create_sheet(chart_sheet_name)
                ws_chart.sheet_view.showGridLines = False # White background
                
                num_series = num_cols - 1
                # Create chart referencing data on 'ws' (Data Sheet)
                if chart_type == 'line':
                    chart = create_line_chart(ws, chart_title, end_row, 1, num_series)
                elif chart_type == 'stacked':
                    chart = create_stacked_chart(ws, chart_title, end_row, 1, num_series)
                elif chart_type == 'scatter':
                    chart = create_scatter_chart(ws, chart_title, end_row, 1, num_series)
                else:
                    chart = create_bar_chart(ws, chart_title, end_row, 1, num_series)
                
                # Create Margins for Visual Centering (Screen)
                # Pushing chart to the right to simulate center on typical screen
                ws_chart.column_dimensions['A'].width = 20 
                ws_chart.row_dimensions[1].height = 20
                
                # Print Settings: True Centering for PDF/Paper
                ws_chart.page_setup.centerHorizontally = True
                ws_chart.page_setup.centerVertically = True
                ws_chart.page_setup.orientation = ws_chart.ORIENTATION_LANDSCAPE
                ws_chart.page_setup.paperSize = ws_chart.PAPERSIZE_A4

                # Chart Size: Balanced
                chart.width = 26 # ~26cm
                chart.height = 16 # ~16cm
                
                ws_chart.add_chart(chart, "B2") 
            except Exception as e:
                import traceback
                debug_log.append(f"Error processing sheet {sheet_data.get('name')}: {str(e)}")
                debug_log.append(traceback.format_exc())

        # Remove empty default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Add Debug Sheet if errors
        if debug_log:
            ws_debug = wb.create_sheet("DEBUG LOG")
            for i, log in enumerate(debug_log):
                ws_debug.cell(row=i+1, column=1, value=log)
                
        # Save to BytesIO
        enforce_global_font(wb)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        return generate_error_excel(f"Global Error: {str(e)}\n\n{traceback.format_exc()}")


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_POST(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            # Use json.loads safely
            try:
                data = json.loads(post_data.decode('utf-8'))
            except json.JSONDecodeError as e:
                self.send_response(200) # Send 200 to bypass fallback
                excel_bytes = generate_error_excel(f"JSON Decode Error: {str(e)}")
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="Error_Report.xlsx"')
                self.end_headers()
                self.wfile.write(excel_bytes)
                return

            filename = data.get('filename', f'Report_{datetime.now().strftime("%Y-%m-%d")}.xlsx')
            
            # Generate Excel (will catch its own errors)
            excel_bytes = generate_excel(data)
            
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', len(excel_bytes))
            self.end_headers()
            self.wfile.write(excel_bytes)
            
        except Exception as e:
            # Fatal error in handler (shouldn't happen with above try/except blocks)
            # Try to send an excel file with error even here
            self.send_response(200) # Force 200
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="Critical_Error.xlsx"')
            
            import traceback
            err_msg = f"Critical Handler Error: {str(e)}\n{traceback.format_exc()}"
            try:
                err_bytes = generate_error_excel(err_msg)
                self.send_header('Content-Length', len(err_bytes))
                self.end_headers()
                self.wfile.write(err_bytes)
            except:
                # If we can't even generate error excel, fallback to text
                self.end_headers()
                self.wfile.write(err_msg.encode())
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps({"status": "ok", "service": "excel-chart-generator"}).encode())
