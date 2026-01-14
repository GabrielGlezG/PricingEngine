# Excel Chart Generator Service
# Generates native Excel charts using openpyxl

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

app = Flask(__name__)
CORS(app)

# Styling constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row styling"""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL
            cell.border = THIN_BORDER


def create_bar_chart(ws, title, data_range, categories_range, start_row, num_series):
    """Create a bar chart"""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Valor"
    chart.x_axis.title = None
    
    data = Reference(ws, min_col=2, min_row=start_row, max_col=1 + num_series, max_row=data_range)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_range)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 10
    
    return chart


def create_line_chart(ws, title, data_range, categories_range, start_row, num_series):
    """Create a line chart"""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Valor"
    chart.x_axis.title = None
    
    data = Reference(ws, min_col=2, min_row=start_row, max_col=1 + num_series, max_row=data_range)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_range)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    
    return chart


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok", "service": "excel-chart-generator"})


@app.route('/generate-excel', methods=['POST', 'OPTIONS'])
def generate_excel():
    """Generate Excel file with native charts"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        filename = data.get('filename', f'Dashboard_Report_{datetime.now().strftime("%Y-%m-%d")}.xlsx')
        sheets = data.get('sheets', [])
        summary = data.get('summary', None)
        models = data.get('models', None)
        currency_symbol = data.get('currencySymbol', '$')
        timezone_offset = data.get('timezoneOffset', -3)  # Default to Chile (UTC-3)
        
        # Generic export parameters
        title = data.get('title', 'REPORTE DE DASHBOARD')
        subtitle = data.get('subtitle', '')
        filters_data = data.get('filters', {})
        
        # Calculate local time with offset
        from datetime import timedelta
        local_time = datetime.utcnow() + timedelta(hours=timezone_offset)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        
        # 1. Create Summary Sheet if provided (Dashboard style)
        if summary:
            ws = wb.create_sheet("Resumen Ejecutivo", 0)
            
            # Title
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Generado: {local_time.strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(italic=True, color="666666")
            
            # Metrics table
            ws['A4'] = "Métrica"
            ws['B4'] = "Valor"
            style_header_row(ws, 4, 2)
            
            metrics = [
                ("Total Modelos", summary.get('total_models', 0)),
                ("Total Marcas", summary.get('total_brands', 0)),
                ("Precio Promedio", summary.get('avg_price', 0)),
                ("Precio Mediano", summary.get('median_price', 0)),
                ("Precio Mínimo", summary.get('min_price', 0)),
                ("Precio Máximo", summary.get('max_price', 0)),
                ("Desviación Estándar", summary.get('price_std_dev', 0)),
                ("Coef. Variación", summary.get('variation_coefficient', 0)),
                ("Descuento Promedio", summary.get('avg_discount_pct', 0)), # New Metric
            ]
            
            for i, (label, value) in enumerate(metrics):
                row = 5 + i
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                if 2 <= i <= 6:  # Currency values
                    ws.cell(row=row, column=2).number_format = f'"{currency_symbol}" #,##0'
                elif i >= 7:  # Percentage (Variation Coeff & Discount)
                    ws.cell(row=row, column=2).number_format = '0.00%'
            
            style_data_rows(ws, 5, 13, 2)
            
            # Filters
            ws['A15'] = "Filtros Aplicados"
            ws['A15'].font = Font(bold=True)
            
            filters = summary.get('filters', {})
            ws['A16'] = "Segmento:"
            ws['B16'] = ', '.join(filters.get('tipoVehiculo', [])) or "Todos"
            ws['A17'] = "Marca:"
            ws['B17'] = ', '.join(filters.get('brand', [])) or "Todas"
            ws['A18'] = "Modelo:"
            ws['B18'] = ', '.join(filters.get('model', [])) or "Todos"

// ... (skipping unchanged code block for sheets loop if needed, but I need to target Models section below) ...

        # 3. Create Models sheet if provided
        if models and len(models) > 0:
            ws = wb.create_sheet("Modelos")
            
            headers = ["Marca", "Modelo", "Versión", "Estado", "Tipo Vehículo", 
                      "Precio c/Bono", "Precio Lista", "Bono", "% Descuento"]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            style_header_row(ws, 1, len(headers))
            
            for row_idx, model in enumerate(models, 2):
                ws.cell(row=row_idx, column=1, value=model.get('brand', ''))
                ws.cell(row=row_idx, column=2, value=model.get('model', ''))
                ws.cell(row=row_idx, column=3, value=model.get('submodel', '-'))
                ws.cell(row=row_idx, column=4, value=model.get('estado', 'N/A'))
                ws.cell(row=row_idx, column=5, value=model.get('tipo_vehiculo', 'N/A'))
                
                precio_bono = model.get('precio_con_bono', 0)
                precio_lista = model.get('precio_lista', 0)
                bono = model.get('bono', 0)
                
                # Calculate Discount % (Bono / Lista)
                if precio_lista and precio_lista > 0:
                    diff = (bono / precio_lista)
                else:
                    diff = 0
                
                ws.cell(row=row_idx, column=6, value=precio_bono).number_format = f'"{currency_symbol}" #,##0'
                ws.cell(row=row_idx, column=7, value=precio_lista).number_format = f'"{currency_symbol}" #,##0'
                ws.cell(row=row_idx, column=8, value=bono).number_format = f'"{currency_symbol}" #,##0'
                ws.cell(row=row_idx, column=9, value=diff).number_format = '0.0%'
            
            end_row = len(models) + 1
            style_data_rows(ws, 2, end_row, len(headers))
            
            # Set column widths
            widths = [15, 20, 25, 12, 15, 18, 18, 15, 15]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
        
        # Remove empty default sheet if it still exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error generating Excel: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
